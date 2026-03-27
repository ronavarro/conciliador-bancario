"""
exporter.py — Genera el archivo Excel de conciliación.
"""

from __future__ import annotations
import io
from typing import List, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from engine import ReconciliationResult


# ── Paleta ──────────────────────────────────────────────────────────────────
C = {
    "red_bg":     "FFC7CE",
    "green_bg":   "C6EFCE",
    "yellow_bg":  "FFEB9C",
    "lblue_bg":   "DAEEF3",
    "gray_bg":    "F2F2F2",
    "white":      "FFFFFF",
    "red_light":  "FFF2F2",
    "green_lt":   "F0FFF0",
    "green_dk":   "EBF1DE",
    "blue_hdr":   "1F4E79",
    "mid_blue":   "2E74B5",
    "dark_red":   "C00000",
    "dark_green": "375623",
    "dark_gray":  "7F7F7F",
}


def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _hdr(cell, bg="1F4E79", fg="FFFFFF", size=10, bold=True):
    cell.font      = Font(name="Arial", bold=bold, color=fg, size=size)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _border()


def _dat(cell, bg="FFFFFF", bold=False, num_fmt=None):
    cell.font      = Font(name="Arial", size=9, bold=bold)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border    = _border()
    if num_fmt:
        cell.number_format = num_fmt


def _write_table(ws, start_row: int, headers: List[str], df: pd.DataFrame,
                 row_bg: str = "FFFFFF", hdr_bg: str = "1F4E79",
                 num_cols: Optional[List[int]] = None) -> int:
    """Escribe encabezados + filas. Devuelve la próxima fila libre."""
    for i, h in enumerate(headers, 1):
        _hdr(ws.cell(row=start_row, column=i, value=h), bg=hdr_bg)

    for r_idx, row in df.iterrows():
        row_num = start_row + 1 + r_idx
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=c_idx, value=val)
            is_num = num_cols and c_idx in num_cols
            _dat(cell, bg=row_bg, num_fmt="#,##0.00" if is_num else None)

    return start_row + 1 + len(df)


def build_excel(result: ReconciliationResult, banco_nombre: str, periodo: str) -> bytes:
    wb = Workbook()

    # ── HOJA 1: RESUMEN ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "RESUMEN"
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 34

    ws.merge_cells("A1:D1")
    ws["A1"] = f"CONCILIACIÓN BANCARIA — {banco_nombre.upper()} — {periodo}"
    ws["A1"].font      = Font(name="Arial", bold=True, size=14, color=C["blue_hdr"])
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill      = PatternFill("solid", fgColor=C["lblue_bg"])
    ws.row_dimensions[1].height = 30

    row = 3
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "ESTADÍSTICAS GENERALES"
    _hdr(ws[f"A{row}"], bg=C["mid_blue"], size=11)
    ws.row_dimensions[row].height = 22

    stats = [
        ("Movimientos en extracto bancario",             result.banco_total,          "", ""),
        ("   Créditos",                                  result.banco_creditos,        f"$ {result.monto_faltantes_creditos + sum(r['Credito'] for _, r in result.faltantes_creditos.iterrows() if False) :,.0f}", ""),
        ("   Débitos",                                   result.banco_debitos,         "", ""),
        ("Asientos en Mayor de Cuentas (período)",       result.mayor_total,           "", ""),
        ("Movimientos CONCILIADOS",                      result.conciliados,           f"{result.pct_conciliado:.1f}%", ""),
        ("   Créditos conciliados",                      result.conciliados_creditos,  "", ""),
        ("   Débitos conciliados",                       result.conciliados_debitos,   "", ""),
    ]
    # Re-calcular importes correctos
    total_cred = result.faltantes_creditos["Credito"].sum() if not result.faltantes_creditos.empty else 0
    total_deb  = result.faltantes_debitos["Debito"].sum()   if not result.faltantes_debitos.empty else 0

    for s in stats:
        row += 1
        ws[f"A{row}"] = s[0]; ws[f"B{row}"] = s[1]
        ws[f"C{row}"] = s[2]; ws[f"D{row}"] = s[3]
        for col in "ABCD":
            _dat(ws[f"{col}{row}"], bg=C["gray_bg"])

    row += 2
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "⚠  MOVIMIENTOS EN BANCO NO REGISTRADOS EN EL SISTEMA"
    _hdr(ws[f"A{row}"], bg=C["dark_red"], size=11)
    ws.row_dimensions[row].height = 22

    total_gi = result.monto_gastos_impuestos
    discrep = [
        ("Créditos sin asiento en mayor",      len(result.faltantes_creditos),   f"$ {total_cred:,.2f}",  '→ Ver "Faltantes Créditos"'),
        ("Débitos sin asiento en mayor",        len(result.faltantes_debitos),    f"$ {total_deb:,.2f}",   '→ Ver "Faltantes Débitos"'),
        ("TOTAL faltantes",                     result.total_faltantes,           f"$ {total_cred + total_deb:,.2f}", ""),
        ("Gastos e impuestos bancarios",        len(result.gastos_impuestos),     f"$ {total_gi:,.2f}",    '→ Ver "Gastos e Impuestos"'),
    ]
    for d in discrep:
        row += 1
        ws[f"A{row}"] = d[0]; ws[f"B{row}"] = d[1]
        ws[f"C{row}"] = d[2]; ws[f"D{row}"] = d[3]
        for col in "ABCD":
            _dat(ws[f"{col}{row}"], bg=C["red_bg"], bold=d[0].startswith("TOTAL"))

    row += 2
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = "ℹ  ASIENTOS EN SISTEMA SIN CORRESPONDENCIA EN BANCO"
    _hdr(ws[f"A{row}"], bg=C["dark_gray"], size=11)
    ws.row_dimensions[row].height = 22

    for label, df_part, col in [
        ("Debe en mayor sin movimiento en banco",  result.mayor_sin_banco_debe,  "Debe"),
        ("Haber en mayor sin movimiento en banco", result.mayor_sin_banco_haber, "Haber"),
    ]:
        row += 1
        amt = df_part[col].sum() if not df_part.empty and col in df_part.columns else 0
        ws[f"A{row}"] = label
        ws[f"B{row}"] = len(df_part)
        ws[f"C{row}"] = f"$ {amt:,.2f}"
        ws[f"D{row}"] = '→ Ver "Mayor sin Banco"'
        for col_ in "ABCD":
            _dat(ws[f"{col_}{row}"], bg=C["yellow_bg"])

    # ── HOJA 2: Faltantes Créditos ────────────────────────────────────────
    ws2 = wb.create_sheet("Faltantes Créditos")
    for col, w in zip("ABCD", [14, 50, 24, 18]):
        ws2.column_dimensions[col].width = w

    ws2.merge_cells("A1:D1")
    ws2["A1"] = f"CRÉDITOS EN BANCO SIN REGISTRO EN SISTEMA — {len(result.faltantes_creditos)} movimiento(s)"
    ws2["A1"].font      = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    ws2["A1"].fill      = PatternFill("solid", fgColor=C["dark_red"])
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 25

    if not result.faltantes_creditos.empty:
        nxt = _write_table(ws2, 2,
            ["Fecha", "Concepto", "Comprobante", "Crédito ($)"],
            result.faltantes_creditos, row_bg=C["red_light"], num_cols=[4])
        ws2.cell(row=nxt, column=3, value="TOTAL")
        ws2.cell(row=nxt, column=4, value=f"=SUM(D3:D{nxt-1})")
        for c in range(1, 5):
            _dat(ws2.cell(row=nxt, column=c), bg=C["red_bg"], bold=True)
        ws2.cell(row=nxt, column=4).number_format = "#,##0.00"

    # ── HOJA 3: Faltantes Débitos ─────────────────────────────────────────
    ws3 = wb.create_sheet("Faltantes Débitos")
    for col, w in zip("ABCD", [14, 50, 24, 18]):
        ws3.column_dimensions[col].width = w

    ws3.merge_cells("A1:D1")
    ws3["A1"] = f"DÉBITOS EN BANCO SIN REGISTRO EN SISTEMA — {len(result.faltantes_debitos)} movimiento(s)"
    ws3["A1"].font      = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    ws3["A1"].fill      = PatternFill("solid", fgColor=C["dark_red"])
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 25

    if not result.faltantes_debitos.empty:
        nxt3 = _write_table(ws3, 2,
            ["Fecha", "Concepto", "Comprobante", "Débito ($)"],
            result.faltantes_debitos, row_bg=C["red_light"], num_cols=[4])
        ws3.cell(row=nxt3, column=3, value="TOTAL")
        ws3.cell(row=nxt3, column=4, value=f"=SUM(D3:D{nxt3-1})")
        for c in range(1, 5):
            _dat(ws3.cell(row=nxt3, column=c), bg=C["red_bg"], bold=True)
        ws3.cell(row=nxt3, column=4).number_format = "#,##0.00"

    # ── HOJA 4: Gastos e Impuestos ────────────────────────────────────────
    ws_gi = wb.create_sheet("Gastos e Impuestos")
    for col, w in zip("ABCDE", [14, 50, 24, 18, 40]):
        ws_gi.column_dimensions[col].width = w

    ws_gi.merge_cells("A1:E1")
    ws_gi["A1"] = f"GASTOS E IMPUESTOS BANCARIOS — {len(result.gastos_impuestos)} movimiento(s)"
    ws_gi["A1"].font      = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    ws_gi["A1"].fill      = PatternFill("solid", fgColor="7030A0")
    ws_gi["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_gi.row_dimensions[1].height = 25

    if not result.gastos_impuestos.empty:
        gi_df = result.gastos_impuestos.copy()
        gi_df["_fecha_dt"] = pd.to_datetime(gi_df["Fecha"], dayfirst=True, errors="coerce")
        gi_df["Mes"] = gi_df["_fecha_dt"].dt.to_period("M").astype(str)

        # Resumen por mes
        ws_gi.merge_cells("A2:E2")
        ws_gi["A2"] = "RESUMEN POR MES"
        _hdr(ws_gi["A2"], bg="7030A0", size=10)

        resumen = (
            gi_df.groupby("Mes")["Debito"]
            .agg(Cantidad="count", Total="sum")
            .reset_index()
        )
        resumen["Total"] = resumen["Total"].abs()
        r_gi = _write_table(ws_gi, 3,
            ["Mes", "Cantidad movimientos", "Total ($)"],
            resumen, row_bg="EDE7F6", hdr_bg="7030A0", num_cols=[3])

        # Total resumen
        ws_gi.cell(row=r_gi, column=2, value="TOTAL")
        ws_gi.cell(row=r_gi, column=3, value=resumen["Total"].sum())
        ws_gi.cell(row=r_gi, column=3).number_format = "#,##0.00"
        for c in range(1, 4):
            _dat(ws_gi.cell(row=r_gi, column=c), bg="D1C4E9", bold=True)

        r_gi += 2
        ws_gi.merge_cells(f"A{r_gi}:E{r_gi}")
        ws_gi[f"A{r_gi}"] = "DETALLE DE MOVIMIENTOS"
        _hdr(ws_gi[f"A{r_gi}"], bg="7030A0", size=10)

        gi_display = gi_df.drop(columns=["_fecha_dt", "Mes"], errors="ignore")
        r_gi += 1
        nxt_gi = _write_table(ws_gi, r_gi,
            ["Fecha", "Concepto", "Comprobante", "Débito ($)", "Crédito ($)"],
            gi_display, row_bg="EDE7F6", hdr_bg="7030A0", num_cols=[4, 5])

        ws_gi.cell(row=nxt_gi, column=3, value="TOTAL")
        ws_gi.cell(row=nxt_gi, column=4, value=result.monto_gastos_impuestos)
        ws_gi.cell(row=nxt_gi, column=4).number_format = "#,##0.00"
        for c in range(1, 6):
            _dat(ws_gi.cell(row=nxt_gi, column=c), bg="D1C4E9", bold=True)

    # ── HOJA 5: Mayor sin Banco ───────────────────────────────────────────
    ws4 = wb.create_sheet("Mayor sin Banco")
    for col, w in zip("ABCDE", [14, 52, 14, 18, 30]):
        ws4.column_dimensions[col].width = w

    ws4.merge_cells("A1:E1")
    ws4["A1"] = "ASIENTOS EN MAYOR SIN CORRESPONDENCIA EN EXTRACTO BANCARIO"
    ws4["A1"].font      = Font(name="Arial", bold=True, size=12, color=C["dark_gray"])
    ws4["A1"].fill      = PatternFill("solid", fgColor=C["yellow_bg"])
    ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 25

    r4 = 2
    if not result.mayor_sin_banco_debe.empty:
        ws4.merge_cells(f"A{r4}:E{r4}")
        ws4[f"A{r4}"] = f"ASIENTOS DEBE — {len(result.mayor_sin_banco_debe)} entradas"
        _hdr(ws4[f"A{r4}"], bg=C["dark_green"], size=10)
        r4 += 1
        r4 = _write_table(ws4, r4,
            ["Fecha", "Descripción ERP", "Asiento", "Debe ($)", "Nota"],
            result.mayor_sin_banco_debe.assign(Nota="En sistema, no en banco"),
            row_bg=C["green_dk"], hdr_bg=C["dark_green"], num_cols=[4])

    if not result.mayor_sin_banco_haber.empty:
        r4 += 1
        ws4.merge_cells(f"A{r4}:E{r4}")
        ws4[f"A{r4}"] = f"ASIENTOS HABER — {len(result.mayor_sin_banco_haber)} entradas"
        _hdr(ws4[f"A{r4}"], bg=C["dark_gray"], size=10)
        r4 += 1
        _write_table(ws4, r4,
            ["Fecha", "Descripción ERP", "Asiento", "Haber ($)", "Nota"],
            result.mayor_sin_banco_haber.assign(Nota="En sistema, no en banco"),
            row_bg=C["gray_bg"], hdr_bg=C["dark_gray"], num_cols=[4])

    # ── HOJA 5: Extracto Completo ─────────────────────────────────────────
    ws5 = wb.create_sheet("Extracto Completo")
    for col, w in zip("ABCDEF", [14, 46, 24, 16, 16, 20]):
        ws5.column_dimensions[col].width = w

    ws5.merge_cells("A1:F1")
    ws5["A1"] = f"EXTRACTO BANCARIO COMPLETO — {banco_nombre} — {result.banco_total} movimientos"
    ws5["A1"].font      = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    ws5["A1"].fill      = PatternFill("solid", fgColor=C["blue_hdr"])
    ws5["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws5.row_dimensions[1].height = 25

    for i, h in enumerate(["Fecha", "Concepto", "Comprobante", "Crédito ($)", "Débito ($)", "Estado"], 1):
        _hdr(ws5.cell(row=2, column=i))

    for r_idx, row_data in result.banco_completo.iterrows():
        rn   = r_idx + 3
        cols = ["Fecha", "Concepto", "Comprobante", "Credito", "Debito", "Estado"]
        vals = [row_data[c] for c in cols]
        is_unc = row_data["Estado"] == "⚠ No en sistema"
        rbg = C["red_bg"] if is_unc else (C["green_lt"] if r_idx % 2 == 0 else C["white"])

        for c_idx, val in enumerate(vals, 1):
            cell = ws5.cell(row=rn, column=c_idx, value=val if val != 0 else None)
            _dat(cell, bg=rbg, num_fmt="#,##0.00" if c_idx in (4, 5) else None)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
